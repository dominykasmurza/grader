"""Answer-key loading and configurable question scoring."""

import os
import re

from .config import *


def normalize_score_map(score_map=None):
    """Return a complete score map for 0..STATEMENTS_PER_QUESTION correct answers."""
    merged = dict(SCORE_BY_CORRECT_COUNT)
    if score_map:
        for key, value in score_map.items():
            merged[int(key)] = float(value)

    normalized = {}
    for i in range(STATEMENTS_PER_QUESTION + 1):
        normalized[i] = float(merged.get(i, 0.0))
    return normalized


def parse_score_map_arg(text):
    """Parse CLI score map, e.g. '0:0,1:0.2,2:0.4,3:0.6,4:1'."""
    if text is None or str(text).strip() == "":
        return normalize_score_map()

    parsed = {}
    for chunk in str(text).split(","):
        chunk = chunk.strip()
        if not chunk:
            continue
        if ":" not in chunk:
            raise ValueError(f"Invalid score map item: {chunk}. Expected format like 2:0.4")
        key_text, value_text = chunk.split(":", 1)
        key = int(key_text.strip())
        value = float(value_text.strip())
        if key < 0 or key > STATEMENTS_PER_QUESTION:
            raise ValueError(f"Score map key must be 0..{STATEMENTS_PER_QUESTION}, got {key}")
        parsed[key] = value

    return normalize_score_map(parsed)


def format_score(value):
    """Compact score formatting for overlays and CSV-friendly console output."""
    try:
        value = float(value)
    except (TypeError, ValueError):
        return ""
    text = f"{value:.2f}"
    text = text.rstrip("0").rstrip(".")
    return text if text else "0"


def normalize_answer(value):
    if value is None:
        return ""
    text = str(value).strip().upper()
    if text in {"TRUE", "T", "1"}:
        return "T"
    if text in {"FALSE", "F", "0"}:
        return "F"
    return ""


def normalize_part_label(part):
    """Normalize a user-defined part label while preserving its display text.

    The historical Part A/B aliases remain supported, but arbitrary labels such
    as ``Form Alpha`` or ``Molecular Biology`` are also valid. ``auto`` and blank
    values mean that no explicit label was supplied.
    """
    if part is None:
        return None

    text = re.sub(r"\s+", " ", str(part)).strip()
    if not text or text.casefold() == "auto":
        return None

    aliases = {
        "A": "A",
        "PART A": "A",
        "THEORY PART A": "A",
        "THEORY_A": "A",
        "THEORY-PART-A": "A",
        "B": "B",
        "PART B": "B",
        "THEORY PART B": "B",
        "THEORY_B": "B",
        "THEORY-PART-B": "B",
    }
    return aliases.get(text.upper(), text)


def normalize_part(part):
    """Backward-compatible alias for :func:`normalize_part_label`."""
    return normalize_part_label(part)


def canonical_part_label(part):
    """Return the case-insensitive lookup key for a part label."""
    normalized = normalize_part_label(part)
    return normalized.casefold() if normalized else None


def _question_number_from_value(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        number = value
    elif isinstance(value, float) and value.is_integer():
        number = int(value)
    else:
        text = str(value).strip()
        match = re.fullmatch(r"\d{1,3}", text) or re.search(r"(\d{1,3})$", text)
        if not match:
            return None
        number = int(match.group(0) if match.lastindex is None else match.group(1))

    if 1 <= number <= NUM_QUESTIONS:
        return number
    return None


def _legacy_part_from_question_code(value):
    text = str(value or "").strip().upper()
    match = re.fullmatch(r"([AB])(\d{1,2})", text)
    return match.group(1) if match else None


def question_code_for(part, question_number):
    part_label = normalize_part_label(part)
    if not part_label:
        return ""
    number = int(question_number)
    if part_label in {"A", "B"}:
        return f"{part_label}{number:02d}"
    return f"{part_label}:{number:02d}"


def _answer_key_entry(part_label, question_number, code, answers):
    return {
        "part_label": normalize_part_label(part_label),
        "question": int(question_number),
        "code": str(code or question_code_for(part_label, question_number)).strip(),
        "answers": list(answers),
    }


def load_answer_key(answer_key_path):
    """Read a generalized Excel answer key.

    Preferred format:

    * column A: question number or question code;
    * column B: part label repeated for every question;
    * columns D-G: the four correct T/F answers.

    The part label may be any non-empty text. It must match a token present in
    the scanned filename or the QR payload when ``--part auto`` is used.

    Legacy rows such as ``A01``/``B50`` in column A remain supported when
    column B is blank.
    """
    if not answer_key_path:
        return {}
    if not os.path.exists(answer_key_path):
        raise FileNotFoundError(f"Answer key not found: {answer_key_path}")

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            "Reading .xlsx answer keys requires openpyxl: pip install openpyxl"
        ) from exc

    wb = load_workbook(answer_key_path, data_only=True, read_only=True)
    ws = wb.active

    answer_key = {}
    for row_number, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        raw_question = row[0] if len(row) > 0 else None
        question_number = _question_number_from_value(raw_question)
        if question_number is None:
            continue  # Header, note, or unrelated row.

        raw_part_label = row[1] if len(row) > 1 else None
        part_label = normalize_part_label(raw_part_label)
        if part_label is None:
            part_label = _legacy_part_from_question_code(raw_question)
        if part_label is None:
            raise ValueError(
                f"Answer-key row {row_number} has question {raw_question!r} but no PartLabel in column B."
            )

        answers = [
            normalize_answer(row[i] if len(row) > i else None)
            for i in range(3, 7)
        ]
        if len(answers) != STATEMENTS_PER_QUESTION or not all(
            answer in {"T", "F"} for answer in answers
        ):
            raise ValueError(
                f"Answer-key row {row_number} ({part_label}, question {question_number}) "
                "must contain four valid T/F answers in columns D-G."
            )

        code_text = str(raw_question).strip() if raw_question is not None else ""
        if re.fullmatch(r"\d+(?:\.0+)?", code_text):
            code_text = question_code_for(part_label, question_number)

        key = (canonical_part_label(part_label), question_number)
        if key in answer_key:
            raise ValueError(
                f"Duplicate answer-key entry for part {part_label!r}, question {question_number}."
            )
        answer_key[key] = _answer_key_entry(
            part_label,
            question_number,
            code_text,
            answers,
        )

    if not answer_key:
        raise ValueError(
            "No valid answer-key rows were found. Expected question in column A, "
            "PartLabel in column B, and T/F answers in columns D-G."
        )
    return answer_key


def answer_key_part_labels(answer_key):
    """Return unique display labels in first-occurrence order."""
    labels = []
    seen = set()

    for key, value in (answer_key or {}).items():
        label = None
        if isinstance(value, dict):
            label = normalize_part_label(value.get("part_label"))
        if label is None and isinstance(key, tuple) and len(key) == 2:
            label = normalize_part_label(key[0])
        if label is None and isinstance(key, str):
            match = re.fullmatch(r"([AB])\d{1,2}", key.strip().upper())
            label = match.group(1) if match else None

        canonical = canonical_part_label(label)
        if canonical and canonical not in seen:
            seen.add(canonical)
            labels.append(label)
    return labels


def answer_key_has_part(answer_key, part_label):
    target = canonical_part_label(part_label)
    return bool(target) and any(
        canonical_part_label(label) == target
        for label in answer_key_part_labels(answer_key)
    )


def _lookup_answer_key_entry(answer_key, part_label, question_number):
    canonical = canonical_part_label(part_label)
    if not canonical:
        return None

    entry = (answer_key or {}).get((canonical, int(question_number)))
    if isinstance(entry, dict):
        return entry

    # Backward compatibility with manually constructed legacy dictionaries.
    legacy_code = question_code_for(part_label, question_number)
    legacy_answers = (answer_key or {}).get(legacy_code)
    if isinstance(legacy_answers, (list, tuple)):
        return _answer_key_entry(
            part_label,
            question_number,
            legacy_code,
            legacy_answers,
        )
    return None


def grade_detected_rows(detected_rows, answer_key, part, score_map=None):
    score_map = normalize_score_map(score_map)
    default_max_question_score = score_map.get(
        STATEMENTS_PER_QUESTION,
        max(score_map.values()) if score_map else float(STATEMENTS_PER_QUESTION),
    )

    graded_rows = []
    metrics = {
        "gradable_cells": 0,
        "correct": 0,
        "incorrect": 0,
        "missing": 0,
        "ambiguous": 0,
        "score": 0.0,
        "max_score": 0.0,
        "linear_correct_cells": 0,
        "question_count": 0,
        "score_map": score_map,
    }

    for q_idx, row in enumerate(detected_rows, start=1):
        entry = _lookup_answer_key_entry(answer_key, part, q_idx)
        code = entry["code"] if entry else question_code_for(part, q_idx)
        correct = entry["answers"] if entry else [""] * STATEMENTS_PER_QUESTION
        result = []
        correct_count = 0
        gradable_count = 0

        for s_idx, detected in enumerate(row):
            correct_answer = correct[s_idx] if s_idx < len(correct) else ""
            if correct_answer in {"T", "F"}:
                metrics["gradable_cells"] += 1
                gradable_count += 1

                if detected == correct_answer:
                    result.append("correct")
                    metrics["correct"] += 1
                    metrics["linear_correct_cells"] += 1
                    correct_count += 1
                elif detected in {"T", "F"}:
                    result.append("incorrect")
                    metrics["incorrect"] += 1
                else:
                    result.append("missing")
                    metrics["missing"] += 1
            else:
                result.append("no_key")

        row_score = float(score_map.get(correct_count, 0.0))
        row_max = (
            float(score_map.get(gradable_count, default_max_question_score))
            if gradable_count
            else 0.0
        )

        metrics["score"] += row_score
        metrics["max_score"] += row_max
        if gradable_count:
            metrics["question_count"] += 1

        graded_rows.append({
            "question": q_idx,
            "code": code,
            "part_label": normalize_part_label(part) or "",
            "detected": row,
            "correct": correct,
            "result": result,
            "correct_count": correct_count,
            "gradable_count": gradable_count,
            "score": row_score,
            "max_score": row_max,
        })

    return graded_rows, metrics
